from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.utils.http import url_has_allowed_host_and_scheme
from django.db.models import Count, Q
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, StreamingHttpResponse
from django.urls import reverse
from django.db import models as db_models
from django.utils import timezone
from django.core.cache import cache
from datetime import timedelta

import after_response

from .models import VozaviForm, Question, VozaviResponse, Answer
from .templates_data import FORM_TEMPLATES
from .activity import log_event


@after_response.enable
def _notify_new_response(form_id, results_url):
    """Envoi de la notification après la réponse HTTP (thread, sans bloquer)."""
    from .tasks import send_new_response_email
    send_new_response_email(form_id, results_url)


# ── DASHBOARD ─────────────────────────────────────────────────────────────────

@login_required
def dashboard(request):
    forms_qs = VozaviForm.objects.filter(user=request.user)
    agg = forms_qs.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(status='active')),
    )
    total_responses = VozaviResponse.objects.filter(form__user=request.user).count()
    recent_forms = forms_qs.annotate(
        nb_responses=Count('responses', distinct=True),
    ).order_by('-updated_at')[:6]

    return render(request, 'app/admin/dashboard.html', {
        'total_forms': agg['total'],
        'total_active': agg['active'],
        'total_responses': total_responses,
        'recent_forms': recent_forms,
    })


@login_required
def account_settings(request):
    """Gestion du compte : e-mail, mot de passe, suppression."""
    from django.contrib.auth import update_session_auth_hash
    from django.contrib.auth.models import User
    from django.core.validators import validate_email
    from django.core.exceptions import ValidationError

    user = request.user
    ok = {}      # messages de succès par section
    errors = {}  # messages d'erreur par section

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'email':
            new_email = request.POST.get('email', '').strip().lower()
            try:
                validate_email(new_email)
            except ValidationError:
                errors['email'] = "Adresse e-mail invalide."
            else:
                if User.objects.filter(email__iexact=new_email).exclude(pk=user.pk).exists():
                    errors['email'] = "Cette adresse e-mail est déjà utilisée."
                else:
                    user.email = new_email
                    user.save(update_fields=['email'])
                    ok['email'] = "Adresse e-mail mise à jour."

        elif action == 'password':
            current = request.POST.get('current_password', '')
            new1 = request.POST.get('new_password1', '')
            new2 = request.POST.get('new_password2', '')
            if not user.check_password(current):
                errors['password'] = "Mot de passe actuel incorrect."
            elif len(new1) < 8:
                errors['password'] = "Le nouveau mot de passe doit contenir au moins 8 caractères."
            elif new1 != new2:
                errors['password'] = "Les deux mots de passe ne correspondent pas."
            else:
                user.set_password(new1)
                user.save()
                update_session_auth_hash(request, user)  # garde la session active
                ok['password'] = "Mot de passe modifié."

    return render(request, 'vozavi/account/settings.html', {'ok': ok, 'errors': errors})


@login_required
def delete_account(request):
    """Suppression définitive du compte et de toutes ses données."""
    user = request.user
    if request.method == 'POST':
        if user.check_password(request.POST.get('password', '')):
            username = user.get_username()
            logout(request)
            user.delete()  # supprime en cascade les formulaires, réponses, etc.
            log_event('account_deleted', label=username)
            return redirect('home')
        return render(request, 'vozavi/account/delete_account.html', {
            'error': "Mot de passe incorrect. Compte non supprimé.",
            'nb_forms': user.vozavi_forms.count(),
        })
    return render(request, 'vozavi/account/delete_account.html', {
        'nb_forms': user.vozavi_forms.count(),
    })


@login_required
def forms_list(request):
    """Liste complète des formulaires avec recherche, filtre statut, tri et pagination."""
    from django.core.paginator import Paginator

    qs = VozaviForm.objects.filter(user=request.user).annotate(
        nb_responses=Count('responses', distinct=True),
    )

    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(brand_name__icontains=q))

    status = request.GET.get('status', '').strip()
    if status in ('draft', 'active', 'closed'):
        qs = qs.filter(status=status)

    sort = request.GET.get('sort', 'recent')
    sort_map = {
        'recent': '-updated_at',
        'created': '-created_at',
        'responses': '-nb_responses',
        'title': 'title',
    }
    qs = qs.order_by(sort_map.get(sort, '-updated_at'))

    paginator = Paginator(qs, 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Conserve les filtres dans les liens de pagination (sans 'page')
    params = request.GET.copy()
    params.pop('page', None)
    querystring = params.urlencode()

    return render(request, 'vozavi/builder/forms_list.html', {
        'page_obj': page_obj,
        'total_count': paginator.count,
        'q': q, 'status': status, 'sort': sort,
        'querystring': querystring,
    })


# ── AUTHENTIFICATION ──────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None
    if request.method == 'POST':
        ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', '')).split(',')[0].strip()
        rate_key = f'login_attempts_{ip}'
        attempts = cache.get(rate_key, 0)
        if attempts >= 5:
            error = "Trop de tentatives. Réessayez dans 15 minutes."
            return render(request, 'account/login.html', {'error': error})

        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_active:
            cache.delete(rate_key)
            login(request, user)
            next_url = request.POST.get('next') or request.GET.get('next')
            if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                next_url = 'dashboard'
            return redirect(next_url)
        else:
            cache.set(rate_key, attempts + 1, 900)
            error = "Identifiant ou mot de passe incorrect."

    return render(request, 'account/login.html', {'error': error})


def logout_view(request):
    if request.method == 'POST':
        logout(request)
    return redirect('login')


def signup_view(request):
    import secrets as _secrets
    from django.contrib.auth.models import User

    if request.user.is_authenticated:
        return redirect('dashboard')

    claim_pk = request.GET.get('claim') or request.POST.get('claim')
    error = None

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password1', '')

        if not username or not email or not password:
            error = "Veuillez renseigner tous les champs."
        elif len(password) < 8:
            error = "Le mot de passe doit contenir au moins 8 caractères."
        elif User.objects.filter(username=username).exists() or User.objects.filter(email=email).exists():
            error = "Ces informations sont déjà utilisées."
        else:
            user = User.objects.create_user(username=username, email=email, password=password)
            login(request, user)
            log_event('user_signup', actor=user, request=request, label=user.get_username())

            # Claim anonymous form — only if it matches the current session
            if claim_pk:
                try:
                    session_pk = request.session.get('guest_form_pk')
                    if session_pk is None or int(claim_pk) != session_pk:
                        raise ValueError('claim mismatch')
                    anon_form = VozaviForm.objects.get(pk=int(claim_pk), user=None)
                    anon_form.user = user
                    if not anon_form.slug:
                        anon_form.slug = _secrets.token_urlsafe(8)
                    anon_form.status = 'active'
                    anon_form.save()
                    # Clear guest session key
                    request.session.pop('guest_form_pk', None)
                    log_event('guest_form_claimed', actor=user, target=anon_form, request=request)
                    return redirect('share_form', pk=anon_form.pk)
                except (VozaviForm.DoesNotExist, ValueError, TypeError):
                    pass

            # Onboarding : inscription fraîche → écran de bienvenue (choix du modèle)
            return redirect(reverse('new_form') + '?bienvenue=1')

    return render(request, 'account/signup.html', {'error': error, 'claim_pk': claim_pk})


# ── HOME ──────────────────────────────────────────────────────────────────────

def home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'vozavi/landing.html')


def custom_404(request, exception=None):
    return render(request, '404.html', status=404)


def og_image_view(request):
    """Génère l'image Open Graph 1200×630 px pour les aperçus WhatsApp/iMessage/etc."""
    import io, os
    from PIL import Image, ImageDraw, ImageFont

    W, H = 1200, 630
    IND   = (79, 70, 184)    # #4F46B8
    IND_D = (32, 24, 110)    # fond sombre
    IND_L = (110, 102, 214)
    AMB   = (242, 169, 59)   # #F2A93B
    WHITE = (255, 255, 255)
    MUTED  = (185, 181, 235)
    SUBTLE = (135, 130, 205)

    img  = Image.new('RGB', (W, H), IND)
    draw = ImageDraw.Draw(img, 'RGBA')

    # ── Orbes décoratifs ──────────────────────────────────────────────
    draw.ellipse((-180, -180, 400, 400), fill=(*IND_D, 110))
    draw.ellipse((820,  -200, 1380, 360), fill=(*IND_L,  35))
    draw.ellipse((940,   380, 1360, 800), fill=(*IND_D,  70))

    # Grille de points (texture)
    for gx in range(0, W, 44):
        for gy in range(0, H, 44):
            draw.ellipse([gx-1, gy-1, gx+1, gy+1], fill=(*WHITE, 14))

    # ── Logo vozavi — bulle + étoile ──────────────────────────────────
    # Le SVG source a viewBox="0 0 64 64".
    # Icône : bulle x=10..54, y=4..44 + queue jusqu'à y=53
    # On la reproduit à l'échelle S, centrée verticalement à gauche.
    S   = 5.6        # facteur d'échelle
    OX  = 110        # décalage horizontal
    OY  = (H - 53 * S) / 2   # centré verticalement

    def px(x): return OX + x * S
    def py(y): return OY + y * S
    def r(v):  return round(v * S)

    # Bulle arrondie : rect x=10..54, y=4..44, radius=10
    draw.rounded_rectangle(
        [px(10), py(4), px(54), py(44)],
        radius=r(10),
        fill=(*IND_D, 230),
    )
    # Queue de la bulle : triangle (21,44)→(21,53)→(30,44)
    draw.polygon(
        [(px(21), py(44)), (px(21), py(53)), (px(30), py(44))],
        fill=(*IND_D, 230),
    )

    # Étoile : coordonnées absolues calculées depuis le path SVG relatif
    # M32 12  l3.4 6.9  7.6 1.1  -5.5 5.4  1.3 7.6  -6.8-3.6  -6.8 3.6  1.3-7.6  -5.5-5.4  7.6-1.1  z
    star_pts_svg = [
        (32.0, 12.0), (35.4, 18.9), (43.0, 20.0), (37.5, 25.4),
        (38.8, 33.0), (32.0, 29.4), (25.2, 33.0), (26.5, 25.4),
        (21.0, 20.0), (28.6, 18.9),
    ]
    draw.polygon([(px(x), py(y)) for x, y in star_pts_svg], fill=AMB)

    # ── Barre accent amber ────────────────────────────────────────────
    bar_x = round(px(54)) + 46
    draw.rectangle([bar_x, 185, bar_x + 5, 460], fill=AMB)

    # ── Texte ─────────────────────────────────────────────────────────
    def load_font(size, bold=False):
        candidates = [
            f"C:/Windows/Fonts/{'calibrib' if bold else 'calibri'}.ttf",
            f"C:/Windows/Fonts/{'arialbd'  if bold else 'arial'}.ttf",
            f"C:/Windows/Fonts/{'verdanab' if bold else 'verdana'}.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans{}.ttf".format('-Bold' if bold else ''),
            "/usr/share/fonts/truetype/liberation/LiberationSans{}.ttf".format('-Bold' if bold else '-Regular'),
            "/System/Library/Fonts/Helvetica.ttc",
        ]
        for p in candidates:
            if os.path.isfile(p):
                try:
                    return ImageFont.truetype(p, size)
                except Exception:
                    continue
        try:
            return ImageFont.load_default(size=size)
        except TypeError:
            return ImageFont.load_default()

    f_title  = load_font(116, bold=True)
    f_tag    = load_font(38)
    f_domain = load_font(28)

    tx = bar_x + 26

    draw.text((tx, 195), 'vozavi',                         font=f_title,  fill=WHITE)
    draw.text((tx, 345), "Formulaires d'avis gratuits,",   font=f_tag,    fill=MUTED)
    draw.text((tx, 395), "prêts en 2 minutes.",             font=f_tag,    fill=MUTED)
    draw.text((tx, 476), 'vozavi.com',                     font=f_domain, fill=SUBTLE)

    buf = io.BytesIO()
    img.save(buf, 'PNG', optimize=True)
    buf.seek(0)

    resp = HttpResponse(buf.read(), content_type='image/png')
    resp['Cache-Control'] = 'public, max-age=86400'
    return resp


def robots_txt(request):
    """robots.txt : autorise l'indexation des pages publiques, bloque l'app privée."""
    host = f"{request.scheme}://{request.get_host()}"
    lines = [
        "User-agent: *",
        "Allow: /$",
        "Disallow: /dashboard/",
        "Disallow: /forms/",
        "Disallow: /new/",
        "Disallow: /connexion/",
        "Disallow: /inscription/",
        "Disallow: /vz-control-panel/",
        "Disallow: /mot-de-passe/",
        "",
        f"Sitemap: {host}/sitemap.xml",
    ]
    return HttpResponse("\n".join(lines) + "\n", content_type="text/plain")


def sitemap_xml(request):
    """Sitemap des pages publiques (pas l'app privée ni les formulaires individuels)."""
    host = f"{request.scheme}://{request.get_host()}"
    names = ['home', 'contact', 'aide', 'blog', 'cgu', 'confidentialite']
    urls = [f"{host}{reverse(n)}" for n in names]
    body = ['<?xml version="1.0" encoding="UTF-8"?>',
            '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for u in urls:
        body.append(f"  <url><loc>{u}</loc></url>")
    body.append('</urlset>')
    return HttpResponse("\n".join(body), content_type="application/xml")


def cgu_view(request):
    return render(request, 'vozavi/cgu.html', {})


def confidentialite_view(request):
    return render(request, 'vozavi/confidentialite.html', {})


def blog_view(request):
    return render(request, 'vozavi/blog.html', {})


def aide_view(request):
    return render(request, 'vozavi/aide.html', {})


def contact_view(request):
    sent = False
    error = None
    if request.method == 'POST':
        # Anti-spam : honeypot rempli → bot. On simule le succès sans rien enregistrer.
        if request.POST.get('website', '').strip():
            return render(request, 'vozavi/contact.html', {'sent': True, 'error': None})

        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        category = request.POST.get('category', '').strip()
        message = request.POST.get('message', '').strip()
        if not name or not email or not message:
            error = "Veuillez remplir tous les champs obligatoires."
        else:
            from .models import ContactMessage
            ContactMessage.objects.create(
                name=name,
                email=email,
                category=category,
                message=message,
            )
            log_event('contact_message', request=request, label=name, category=category)
            sent = True
    return render(request, 'vozavi/contact.html', {'sent': sent, 'error': error})


# ── VOZAVI FORM BUILDER ────────────────────────────────────────────────────────

def _guest_form_or_404(request, pk):
    """Return the VozaviForm if the requester (logged-in or guest session) owns it."""
    from django.http import Http404
    vform = get_object_or_404(VozaviForm, pk=pk)
    if request.user.is_authenticated:
        if vform.user_id != request.user.pk:
            raise Http404
    else:
        if request.session.get('guest_form_pk') != pk:
            raise Http404
    return vform


def _cleanup_old_guest_forms():
    """Delete unclaimed guest forms older than 48 h."""
    cutoff = timezone.now() - timedelta(hours=48)
    VozaviForm.objects.filter(user=None, created_at__lt=cutoff).delete()


def new_form(request):
    if request.method == 'POST':
        template_key = request.POST.get('template_key', 'scratch')
        tpl = FORM_TEMPLATES.get(template_key, FORM_TEMPLATES['scratch'])
        user = request.user if request.user.is_authenticated else None
        vform = VozaviForm.objects.create(
            user=user,
            title=tpl['title'],
            template_key=template_key,
            status='draft',
        )
        Question.objects.bulk_create([
            Question(form=vform, type=q['type'], label=q['label'],
                     required=q['required'], position=i, options=q['options'])
            for i, q in enumerate(tpl['questions'])
        ])
        log_event('form_created', actor=user, target=vform, request=request)
        if user is None:
            request.session['guest_form_pk'] = vform.pk
            try:
                from .tasks import cleanup_old_guest_forms_task
                # delay_on_commit : n'envoie la tâche qu'après commit transaction DB
                cleanup_old_guest_forms_task.delay_on_commit()
            except Exception:
                pass  # Celery absent — cleanup ignoré, pas de blocage requête
        return redirect('edit_form', pk=vform.pk)
    return render(request, 'vozavi/builder/new.html', {
        'welcome': request.GET.get('bienvenue') == '1',
    })


def edit_form(request, pk):
    vform = _guest_form_or_404(request, pk)
    questions = vform.questions.order_by('position')
    is_guest = not request.user.is_authenticated
    return render(request, 'vozavi/builder/edit.html', {
        'vform': vform, 'questions': questions, 'is_guest': is_guest,
    })


def update_form_meta(request, pk):
    import re as _re
    vform = _guest_form_or_404(request, pk)
    if request.method == 'POST':
        vform.title = request.POST.get('title', vform.title).strip() or vform.title
        vform.description = request.POST.get('description', vform.description).strip()
        vform.brand_name = request.POST.get('brand_name', vform.brand_name).strip()
        brand_color = request.POST.get('brand_color', vform.brand_color)
        if _re.fullmatch(r'#[0-9A-Fa-f]{6}', brand_color):
            vform.brand_color = brand_color
        if 'logo' in request.FILES:
            logo_file = request.FILES['logo']
            try:
                from PIL import Image as _Image
                img = _Image.open(logo_file)
                img.verify()
                logo_file.seek(0)
                if img.format in ('JPEG', 'PNG', 'WEBP', 'GIF'):
                    vform.logo = logo_file
            except Exception:
                pass
        vform.save()
        return HttpResponse('<span class="save-ok">Enregistré ✓</span>')
    return HttpResponse(status=204)


@login_required
def update_form_notify(request, pk):
    """Active/désactive les notifications e-mail de nouvelles réponses (HTMX)."""
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if request.method == 'POST':
        vform.notify_responses = request.POST.get('notify_responses') == 'on'
        vform.save(update_fields=['notify_responses', 'updated_at'])
        return HttpResponse('<span class="save-ok">Enregistré ✓</span>')
    return HttpResponse(status=204)


def add_question(request, pk):
    vform = _guest_form_or_404(request, pk)
    if request.method == 'POST':
        q_type = request.POST.get('type', 'text')
        defaults = {
            'rating': ('Quelle note donneriez-vous ?', {'max': 5}),
            'single_choice': ('Choisissez une option', {'choices': ['Option 1', 'Option 2', 'Option 3']}),
            'multiple_choice': ("Sélectionnez tout ce qui s'applique", {'choices': ['Option 1', 'Option 2', 'Option 3']}),
            'text': ('Votre commentaire', {}),
            'grid': ('Évaluez les critères suivants', {'criteria': ['Critère 1', 'Critère 2', 'Critère 3'], 'max': 5}),
            'contact': ('Vos coordonnées', {'fields': ['first_name', 'email']}),
        }
        max_pos = vform.questions.aggregate(m=db_models.Max('position'))['m']
        label, options = defaults.get(q_type, ('Nouvelle question', {}))
        Question.objects.create(
            form=vform, type=q_type, label=label,
            required=False, position=(max_pos or 0) + 1, options=options,
        )
        # Repli sans JS : si la requête n'est pas HTMX (le <form> natif a été
        # soumis car HTMX ne s'est pas chargé), on recharge l'éditeur au lieu
        # de renvoyer un fragment de page nu.
        if request.headers.get('HX-Request') != 'true':
            return redirect('edit_form', pk=vform.pk)
        questions = vform.questions.order_by('position')
        return render(request, 'vozavi/builder/partials/questions_list.html', {'vform': vform, 'questions': questions})
    return HttpResponse(status=405)


def update_question(request, pk, qid):
    vform = _guest_form_or_404(request, pk)
    q = get_object_or_404(Question, pk=qid, form=vform)
    if request.method == 'POST':
        label = request.POST.get('label', '').strip()
        if label:
            q.label = label
        q.required = request.POST.get('required') in ('on', 'true', '1')
        if q.type in ('single_choice', 'multiple_choice'):
            choices = [c.strip() for c in request.POST.getlist('choices') if c.strip()]
            if choices:
                q.options = {'choices': choices}
        elif q.type == 'grid':
            criteria = [c.strip() for c in request.POST.getlist('criteria') if c.strip()]
            if criteria:
                q.options = {**q.options, 'criteria': criteria}
        elif q.type == 'contact':
            _valid_keys = {'first_name', 'last_name', 'email', 'phone', 'company', 'job_title'}
            fields = [f for f in request.POST.getlist('contact_fields') if f in _valid_keys]
            q.options = {'fields': fields or ['email']}
        q.save()
        return render(request, 'vozavi/builder/partials/question_card.html', {'vform': vform, 'q': q})
    return HttpResponse(status=405)


def move_question(request, pk, qid):
    vform = _guest_form_or_404(request, pk)
    q = get_object_or_404(Question, pk=qid, form=vform)
    if request.method == 'POST':
        direction = request.POST.get('direction', 'down')
        qs = list(vform.questions.order_by('position'))
        idx = next((i for i, x in enumerate(qs) if x.pk == q.pk), None)
        if idx is not None:
            if direction == 'up' and idx > 0:
                qs[idx].position, qs[idx - 1].position = qs[idx - 1].position, qs[idx].position
                qs[idx].save()
                qs[idx - 1].save()
            elif direction == 'down' and idx < len(qs) - 1:
                qs[idx].position, qs[idx + 1].position = qs[idx + 1].position, qs[idx].position
                qs[idx].save()
                qs[idx + 1].save()
        questions = vform.questions.order_by('position')
        return render(request, 'vozavi/builder/partials/questions_list.html', {'vform': vform, 'questions': questions})
    return HttpResponse(status=405)


def delete_question(request, pk, qid):
    vform = _guest_form_or_404(request, pk)
    q = get_object_or_404(Question, pk=qid, form=vform)
    if request.method == 'POST':
        q.delete()
        return HttpResponse('')
    return HttpResponse(status=405)


@login_required
def publish_form(request, pk):
    import secrets as _secrets
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if request.method == 'POST':
        if not vform.slug:
            vform.slug = _secrets.token_urlsafe(8)
        vform.status = 'active'
        vform.save()
        log_event('form_published', actor=request.user, target=vform, request=request)
        return redirect('share_form', pk=vform.pk)
    return HttpResponse(status=405)


@login_required
def toggle_form_status(request, pk):
    """Ferme un formulaire actif, ou rouvre un formulaire fermé."""
    import secrets as _secrets
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if request.method != 'POST':
        return HttpResponse(status=405)
    if vform.status == 'active':
        vform.status = 'closed'
        vform.save()
        log_event('form_closed', actor=request.user, target=vform, request=request)
    elif vform.status == 'closed':
        if not vform.slug:
            vform.slug = _secrets.token_urlsafe(8)
        vform.status = 'active'
        vform.save()
        log_event('form_reopened', actor=request.user, target=vform, request=request)
    # Un brouillon doit d'abord être publié : on le laisse inchangé.
    next_url = request.POST.get('next')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('forms_list')


@login_required
def duplicate_form(request, pk):
    """Crée une copie (brouillon) d'un formulaire existant avec ses questions."""
    if request.method != 'POST':
        return HttpResponse(status=405)
    original = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    questions = list(original.questions.order_by('position'))

    copy = VozaviForm.objects.create(
        user=request.user,
        title=f"{original.title} (copie)",
        template_key=original.template_key,
        brand_name=original.brand_name,
        description=original.description,
        brand_color=original.brand_color,
        is_anonymous=original.is_anonymous,
        notify_responses=original.notify_responses,
        status='draft',  # la copie démarre en brouillon, sans slug ni réponses
    )
    Question.objects.bulk_create([
        Question(form=copy, type=q.type, label=q.label,
                 required=q.required, position=q.position, options=q.options)
        for q in questions
    ])
    log_event('form_duplicated', actor=request.user, target=copy, request=request,
              source_id=original.pk)
    return redirect('edit_form', pk=copy.pk)


@login_required
def delete_form(request, pk):
    """Supprime définitivement un formulaire et toutes ses réponses (droit à l'effacement)."""
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if request.method == 'POST':
        cache.delete(f'form_results_{vform.pk}')
        title = vform.title
        nb = vform.responses.count()
        vform.delete()
        log_event('form_deleted', actor=request.user, label=title, request=request,
                  responses=nb)
        return redirect('forms_list')
    return render(request, 'vozavi/builder/delete_form.html', {
        'vform': vform,
        'nb_responses': vform.responses.count(),
    })


@login_required
def share_form(request, pk):
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if vform.status == 'draft':
        return redirect('edit_form', pk=vform.pk)
    public_url = request.build_absolute_uri(reverse('public_form', args=[vform.slug]))
    brand = vform.brand_name or vform.title
    share_text = f"Donnez votre avis sur {brand} 👉 {public_url}"
    return render(request, 'vozavi/builder/share.html', {
        'vform': vform, 'public_url': public_url,
        'brand': brand, 'share_text': share_text,
    })


@login_required
def qr_code_view(request, pk):
    import qrcode
    import io
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    if not vform.slug:
        return HttpResponse(status=404)
    public_url = request.build_absolute_uri(reverse('public_form', args=[vform.slug]))
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(public_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#2C2C3A', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='image/png')


def preview_form(request, pk):
    vform = _guest_form_or_404(request, pk)
    questions = vform.questions.order_by('position')
    return render(request, 'vozavi/public/form.html', {'vform': vform, 'questions': questions, 'preview': True})


def public_form(request, slug):
    vform = get_object_or_404(VozaviForm, slug=slug, status='active')
    questions = list(vform.questions.order_by('position'))
    errors = {}

    if request.method == 'POST':
        # Anti-spam : si le honeypot est rempli, c'est un bot. On simule le succès
        # (redirection vers le remerciement) sans rien enregistrer.
        if request.POST.get('website', '').strip():
            return redirect('public_form_thanks', slug=slug)

        # Phase 1 — validate without touching the DB
        collected = {}
        for q in questions:
            field = f'q_{q.pk}'
            if q.type == 'multiple_choice':
                value = request.POST.getlist(field)
            elif q.type == 'rating':
                raw = request.POST.get(field, '')
                value = int(raw) if raw.isdigit() else 0
            elif q.type == 'contact':
                enabled = q.options.get('fields', ['email'])
                value = {k: request.POST.get(f'{field}_{k}', '').strip() for k in enabled}
                value = {k: v for k, v in value.items() if v}
            else:
                value = request.POST.get(field, '').strip()
            if q.required and not value:
                errors[str(q.pk)] = 'Ce champ est obligatoire.'
            collected[q.pk] = value

        # Phase 2 — write only if all fields pass
        if not errors:
            response = VozaviResponse.objects.create(form=vform)
            Answer.objects.bulk_create([
                Answer(response=response, question_id=qid, value=val)
                for qid, val in collected.items()
            ])
            cache.delete(f'form_results_{vform.pk}')
            log_event('response_received', actor=vform.user, target=vform, request=request)

            # Notifie le créateur en arrière-plan (thread, après la réponse HTTP)
            if vform.notify_responses and vform.user_id and vform.user.email:
                results_url = request.build_absolute_uri(
                    reverse('form_results', args=[vform.pk]))
                _notify_new_response.after_response(vform.pk, results_url)

            return redirect('public_form_thanks', slug=slug)

    return render(request, 'vozavi/public/form.html', {'vform': vform, 'questions': questions, 'errors': errors})


def public_form_thanks(request, slug):
    vform = get_object_or_404(VozaviForm, slug=slug)
    return render(request, 'vozavi/public/thanks.html', {'vform': vform})


def form_og_image_view(request, slug):
    """Image Open Graph 1200×630 personnalisée par formulaire (aperçu WhatsApp/iMessage)."""
    import io, os, re
    from PIL import Image, ImageDraw, ImageFont

    vform = get_object_or_404(VozaviForm, slug=slug, status='active')

    W, H = 1200, 630
    WHITE = (255, 255, 255)

    def hex_to_rgb(h):
        h = (h or '').strip().lstrip('#')
        if not re.fullmatch(r'[0-9a-fA-F]{6}', h):
            h = '4F46B8'
        return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))

    def mix(c1, c2, t):
        return tuple(round(a + (b - a) * t) for a, b in zip(c1, c2))

    brand = hex_to_rgb(vform.brand_color)
    bg = mix(brand, (0, 0, 0), 0.72)        # fond sombre dérivé de la marque
    accent = mix(brand, (255, 255, 255), 0.12)
    muted = mix(WHITE, bg, 0.42)

    img = Image.new('RGB', (W, H), bg)
    draw = ImageDraw.Draw(img, 'RGBA')

    # Orbes + texture
    draw.ellipse((-180, -180, 400, 400), fill=(*mix(brand, (255, 255, 255), 0.10), 40))
    draw.ellipse((860, 300, 1380, 820), fill=(*mix(brand, (0, 0, 0), 0.4), 90))
    for gx in range(0, W, 44):
        for gy in range(0, H, 44):
            draw.ellipse([gx - 1, gy - 1, gx + 1, gy + 1], fill=(*WHITE, 12))

    # Barre accent (bord gauche)
    draw.rectangle([0, 0, 10, H], fill=accent)

    def load_font(size, bold=False):
        candidates = [
            f"C:/Windows/Fonts/{'calibrib' if bold else 'calibri'}.ttf",
            f"C:/Windows/Fonts/{'arialbd' if bold else 'arial'}.ttf",
            f"C:/Windows/Fonts/{'verdanab' if bold else 'verdana'}.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans{}.ttf".format('-Bold' if bold else ''),
            "/usr/share/fonts/truetype/liberation/LiberationSans{}.ttf".format('-Bold' if bold else '-Regular'),
            "/System/Library/Fonts/Helvetica.ttc",
        ]
        for p in candidates:
            if os.path.isfile(p):
                try:
                    return ImageFont.truetype(p, size)
                except Exception:
                    continue
        try:
            return ImageFont.load_default(size=size)
        except TypeError:
            return ImageFont.load_default()

    LEFT = 92
    top = 120

    # Logo de la marque si présent, sinon pastille étoile
    logo_img = None
    if vform.logo:
        try:
            with vform.logo.open('rb') as fh:
                logo_img = Image.open(io.BytesIO(fh.read())).convert('RGBA')
        except Exception:
            logo_img = None

    if logo_img is not None:
        box = 132
        logo_img.thumbnail((box, box), Image.LANCZOS)
        img.paste(logo_img, (LEFT, top), logo_img)
    else:
        # Pastille arrondie + étoile (logo vozavi)
        b = 128
        draw.rounded_rectangle([LEFT, top, LEFT + b, top + b], radius=28, fill=(*mix(brand, (0, 0, 0), 0.35), 255))
        cx, cy, R, r = LEFT + b / 2, top + b / 2, 46, 19
        import math
        star = []
        for i in range(10):
            ang = -math.pi / 2 + i * math.pi / 5
            rad = R if i % 2 == 0 else r
            star.append((cx + rad * math.cos(ang), cy + rad * math.sin(ang)))
        draw.polygon(star, fill=accent)

    # Eyebrow : nom de marque
    f_eyebrow = load_font(34, bold=True)
    f_title = load_font(82, bold=True)
    f_foot = load_font(32)

    y = top + 168
    brand_name = (vform.brand_name or '').strip()
    if brand_name:
        draw.text((LEFT, y), brand_name.upper()[:42], font=f_eyebrow, fill=accent)
        y += 56

    # Titre (retour à la ligne automatique, max 3 lignes)
    def wrap(text, font, max_w):
        words = text.split()
        lines, cur = [], ''
        for w in words:
            test = (cur + ' ' + w).strip()
            if draw.textlength(test, font=font) <= max_w or not cur:
                cur = test
            else:
                lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines

    title = (vform.title or 'Donnez votre avis').strip()
    lines = wrap(title, f_title, W - LEFT - 80)
    if len(lines) > 3:
        lines = lines[:3]
        lines[-1] = lines[-1].rstrip('…') + '…'
    for ln in lines:
        draw.text((LEFT, y), ln, font=f_title, fill=WHITE)
        y += 96

    # Pied : invitation + domaine
    draw.text((LEFT, H - 88), 'Donnez votre avis en 2 minutes · vozavi.com', font=f_foot, fill=muted)

    buf = io.BytesIO()
    img.save(buf, 'PNG', optimize=True)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='image/png')
    resp['Cache-Control'] = 'public, max-age=86400'
    return resp


# ── DEMO ──────────────────────────────────────────────────────────────────────

def demo_form(request):
    if request.method == 'POST':
        return redirect('demo_form_thanks')
    return render(request, 'vozavi/public/demo.html')


def demo_form_thanks(request):
    return render(request, 'vozavi/public/demo_thanks.html')


# ── VOZAVI RESULTS ─────────────────────────────────────────────────────────────

def _compute_question_stats(q, values):
    if q.type == 'rating':
        mx = q.options.get('max', 5)
        nums = []
        for v in values:
            try:
                n = int(v)
                if 1 <= n <= mx:
                    nums.append(n)
            except (TypeError, ValueError):
                pass
        total = len(nums)
        if not total:
            return {'type': 'rating', 'count': 0, 'avg': None, 'dist': [], 'max': mx, 'avg_pct': 0, 'label': '—'}
        avg = round(sum(nums) / total, 1)
        dist = []
        for star in range(mx, 0, -1):
            cnt = nums.count(star)
            dist.append({'star': star, 'count': cnt, 'pct': round(cnt / total * 100)})
        if avg >= 4.5:
            qual = 'Excellent'
        elif avg >= 4.0:
            qual = 'Très bien'
        elif avg >= 3.0:
            qual = 'Bien'
        elif avg >= 2.0:
            qual = 'Passable'
        else:
            qual = 'À améliorer'
        return {'type': 'rating', 'count': total, 'avg': avg, 'avg_pct': round(avg / mx * 100), 'dist': dist, 'max': mx, 'label': qual}

    elif q.type in ('single_choice', 'multiple_choice'):
        choices = q.options.get('choices', [])
        counts_map = {c: 0 for c in choices}
        total_votes = 0
        for v in values:
            if q.type == 'multiple_choice' and isinstance(v, list):
                for item in v:
                    counts_map[item] = counts_map.get(item, 0) + 1
                    total_votes += 1
            elif isinstance(v, str) and v:
                counts_map[v] = counts_map.get(v, 0) + 1
                total_votes += 1
        choices_data = sorted(
            [{'label': c, 'count': counts_map.get(c, 0),
              'pct': round(counts_map.get(c, 0) / total_votes * 100) if total_votes else 0}
             for c in choices],
            key=lambda x: x['count'], reverse=True,
        )
        return {'type': q.type, 'count': len([v for v in values if v]), 'choices': choices_data, 'total_votes': total_votes}

    elif q.type == 'text':
        texts = [str(v).strip() for v in values if v and str(v).strip()]
        return {'type': 'text', 'count': len(texts), 'texts': texts}

    elif q.type == 'grid':
        criteria = q.options.get('criteria', [])
        texts = [str(v) for v in values if v]
        return {'type': 'grid', 'count': len(texts), 'criteria': criteria, 'texts': texts}

    elif q.type == 'contact':
        enabled = q.options.get('fields', ['email'])
        rows = []
        for v in values:
            if isinstance(v, dict) and v:
                rows.append(v)
        return {'type': 'contact', 'count': len(rows), 'fields': enabled, 'rows': rows}

    return {'type': q.type, 'count': 0}


def _fmt_answer(value):
    if isinstance(value, list):
        return ', '.join(str(v) for v in value)
    return str(value) if value is not None else ''


@login_required
def form_results(request, pk):
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    questions = list(vform.questions.order_by('position'))

    cache_key = f'form_results_{pk}'
    cached = cache.get(cache_key)

    if cached is None:
        # One query for all answers — eliminates N+1
        answers_by_q = {}
        for a in Answer.objects.filter(question__form=vform).values('question_id', 'value'):
            answers_by_q.setdefault(a['question_id'], []).append(a['value'])

        stats_by_q = {}
        avg_sum = 0.0
        avg_weight = 0
        for q in questions:
            stats = _compute_question_stats(q, answers_by_q.get(q.pk, []))
            stats_by_q[q.pk] = stats
            if stats['type'] == 'rating' and stats.get('avg'):
                avg_sum += stats['avg'] * stats['count']
                avg_weight += stats['count']

        global_avg = round(avg_sum / avg_weight, 1) if avg_weight else None

        week_ago = timezone.now() - timedelta(days=7)
        neg_values = Answer.objects.filter(
            question__form=vform,
            question__type='rating',
            response__created_at__gte=week_ago,
        ).values_list('value', flat=True)
        negative_count = sum(1 for v in neg_values if isinstance(v, (int, float)) and v <= 2)

        # Tendance : activité quotidienne sur 14 jours + semaine vs semaine précédente
        dates = [timezone.localdate() - timedelta(days=i) for i in range(13, -1, -1)]
        day_counts = {d: 0 for d in dates}
        since = (timezone.now() - timedelta(days=14)).replace(hour=0, minute=0, second=0, microsecond=0)
        for dt in vform.responses.filter(created_at__gte=since).values_list('created_at', flat=True):
            d = timezone.localtime(dt).date()
            if d in day_counts:
                day_counts[d] += 1
        peak = max(day_counts.values()) or 1
        daily_series = [
            {'label': d.strftime('%d/%m'),
             'count': day_counts[d],
             'pct': round(day_counts[d] / peak * 100)}
            for d in dates
        ]
        last_7 = sum(day_counts[d] for d in dates[-7:])
        prev_7 = sum(day_counts[d] for d in dates[:7])
        if prev_7:
            trend_pct = round((last_7 - prev_7) / prev_7 * 100)
        else:
            trend_pct = None  # pas de base de comparaison

        # Strip PII rows before caching — contact field data must not persist in Redis
        safe_stats = {}
        for qpk, stats in stats_by_q.items():
            if stats.get('type') == 'contact':
                safe_stats[qpk] = {k: v for k, v in stats.items() if k != 'rows'}
            else:
                safe_stats[qpk] = stats

        cached = {
            'total': vform.responses.count(),
            'global_avg': global_avg,
            'negative_count': negative_count,
            'last_response': vform.responses.order_by('-created_at').values_list('created_at', flat=True).first(),
            'stats_by_q': safe_stats,
            'daily_series': daily_series,
            'last_7': last_7,
            'prev_7': prev_7,
            'trend_pct': trend_pct,
        }
        cache.set(cache_key, cached, 120)

    global_avg = cached['global_avg']

    # Reload contact rows fresh from DB (not cached — PII)
    contact_rows = {}
    contact_qs = [q for q in questions if q.type == 'contact']
    if contact_qs:
        for a in Answer.objects.filter(question__in=contact_qs).values('question_id', 'value'):
            contact_rows.setdefault(a['question_id'], []).append(a['value'])

    questions_data = []
    for q in questions:
        stats = cached['stats_by_q'].get(q.pk, {'type': q.type, 'count': 0})
        if q.type == 'contact':
            rows = [v for v in contact_rows.get(q.pk, []) if isinstance(v, dict) and v]
            stats = {**stats, 'rows': rows, 'count': len(rows)}
        questions_data.append({'q': q, 'stats': stats})

    return render(request, 'vozavi/builder/results.html', {
        'vform': vform,
        'total': cached['total'],
        'global_avg': global_avg,
        'global_avg_pct': round(global_avg / 5 * 100) if global_avg else 0,
        'questions_data': questions_data,
        'negative_count': cached['negative_count'],
        'last_response': cached['last_response'],
        'daily_series': cached.get('daily_series', []),
        'last_7': cached.get('last_7', 0),
        'prev_7': cached.get('prev_7', 0),
        'trend_pct': cached.get('trend_pct'),
    })


@login_required
def form_responses(request, pk):
    from django.core.paginator import Paginator

    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    questions = list(vform.questions.order_by('position'))
    total = vform.responses.count()

    all_responses_qs = (
        vform.responses
        .prefetch_related('answers__question')
        .order_by('-created_at')
    )
    paginator = Paginator(all_responses_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    response_rows = []
    for r in page_obj:
        ans_map = {a.question_id: a.value for a in r.answers.all()}
        cells = []
        has_negative = False
        for q in questions:
            raw = ans_map.get(q.pk)
            if raw is None:
                cells.append({'label': q.label, 'display': '—', 'type': q.type, 'raw': None})
            elif q.type == 'rating':
                try:
                    n = int(raw)
                    mx = q.options.get('max', 5)
                    if n <= 2:
                        has_negative = True
                    cells.append({'label': q.label, 'display': n, 'type': 'rating',
                                  'raw': n, 'max': mx, 'pct': round(n / mx * 100)})
                except (TypeError, ValueError):
                    cells.append({'label': q.label, 'display': '—', 'type': q.type, 'raw': None})
            else:
                cells.append({'label': q.label, 'display': _fmt_answer(raw), 'type': q.type, 'raw': raw})
        response_rows.append({'response': r, 'cells': cells, 'has_negative': has_negative})

    return render(request, 'vozavi/builder/responses.html', {
        'vform': vform,
        'total': total,
        'questions': questions,
        'page_obj': page_obj,
        'response_rows': response_rows,
    })


@login_required
def export_results_csv(request, pk):
    import csv as _csv
    import io as _io
    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    questions = list(vform.questions.order_by('position'))

    # Single query — all answers as lightweight dicts, grouped by response
    ans_index = {}
    for a in Answer.objects.filter(question__form=vform).values('response_id', 'question_id', 'value'):
        ans_index.setdefault(a['response_id'], {})[a['question_id']] = a['value']

    def stream():
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(['Date'] + [q.label for q in questions])
        yield buf.getvalue().encode('utf-8-sig')
        for r in vform.responses.order_by('-created_at').iterator(chunk_size=500):
            buf.seek(0)
            buf.truncate(0)
            ans_map = ans_index.get(r.pk, {})
            row = [r.created_at.strftime('%d/%m/%Y %H:%M')]
            for q in questions:
                row.append(_fmt_answer(ans_map.get(q.pk, '')))
            w.writerow(row)
            yield buf.getvalue().encode('utf-8')

    resp = StreamingHttpResponse(stream(), content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = f'attachment; filename="resultats_{vform.slug or vform.pk}.csv"'
    return resp


@login_required
def export_results_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    vform = get_object_or_404(VozaviForm, pk=pk, user=request.user)
    questions = list(vform.questions.order_by('position'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Résultats'

    hdr_fill = PatternFill(start_color='4F46B8', end_color='4F46B8', fill_type='solid')
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill(start_color='F4F3FB', end_color='F4F3FB', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['Date'] + [q.label for q in questions]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
    ws.row_dimensions[1].height = 32

    # Single query — all answers as lightweight dicts, grouped by response
    ans_index = {}
    for a in Answer.objects.filter(question__form=vform).values('response_id', 'question_id', 'value'):
        ans_index.setdefault(a['response_id'], {})[a['question_id']] = a['value']

    for row_idx, r in enumerate(
        vform.responses.order_by('-created_at').iterator(chunk_size=200), 2
    ):
        ans_map = ans_index.get(r.pk, {})
        row_data = [r.created_at.strftime('%d/%m/%Y %H:%M')]
        for q in questions:
            row_data.append(_fmt_answer(ans_map.get(q.pk, '')))
        ws.append(row_data)
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = alt_fill
        ws.cell(row_idx, 1).alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(headers) + 1):
        col_vals = [ws.cell(r, col_idx).value or '' for r in range(1, ws.max_row + 1)]
        best = max((len(str(v)) for v in col_vals), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(best + 3, 14), 52)

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="resultats_{vform.slug or vform.pk}.xlsx"'
    wb.save(resp)
    return resp
