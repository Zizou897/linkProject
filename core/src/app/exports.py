import io
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from django.db.models import Count, Q

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from .models import SessionPresentielle, Formation, Avis


# ── HELPERS EXCEL ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def _style_header_row(ws, row, nb_cols):
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def _style_data_row(ws, row, nb_cols, alternate=False):
    for col in range(1, nb_cols + 1):
        cell = ws.cell(row=row, column=col)
        if alternate:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def _auto_width(ws, min_w=12, max_w=50):
    for col in ws.columns:
        length = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def _excel_response(filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── HELPERS PDF ───────────────────────────────────────────────────────────────

PRIMARY = colors.HexColor('#1F4E79')
LIGHT_BLUE = colors.HexColor('#EBF3FB')

def _pdf_response(filename):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle('titre', fontSize=16, fontName='Helvetica-Bold',
                               textColor=PRIMARY, spaceAfter=6))
    styles.add(ParagraphStyle('sous_titre', fontSize=11, fontName='Helvetica',
                               textColor=colors.grey, spaceAfter=12))
    styles.add(ParagraphStyle('section', fontSize=12, fontName='Helvetica-Bold',
                               textColor=PRIMARY, spaceBefore=14, spaceAfter=6))
    styles.add(ParagraphStyle('corps', fontSize=9, fontName='Helvetica', spaceAfter=4))
    return styles


# ── EXPORT 1 : Avis Excel (global) ────────────────────────────────────────────


# ── EXPORT 2 : Avis Excel ─────────────────────────────────────────────────────

@staff_member_required
def export_avis_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avis"
    ws.row_dimensions[1].height = 30

    headers = [
        'Formation', 'Session (date)', 'Lieu',
        'Nom', 'Prénom', 'Structure', 'Fonction', 'Email', 'Téléphone',
        'Zone géographique', 'Motivations', 'Problème professionnel',
        'Compétences attendues', 'Contraintes calendrier',
        'Contrainte mobilité', 'Format préféré', 'Observation',
        'Date de soumission',
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    qs = Avis.objects.select_related('session__formation').order_by(
        'session__formation__libelle', 'session__date', 'nom'
    )
    for i, avis in enumerate(qs, start=2):
        ws.append([
            avis.session.formation.libelle,
            avis.session.date.strftime('%d/%m/%Y'),
            avis.session.lieu,
            avis.nom, avis.prenom, avis.structure, avis.fonction,
            avis.email, avis.telephone, avis.zone_geographique,
            avis.motivations, avis.probleme_professionnel,
            avis.competences_attendues, avis.contraintes_calendrier,
            'Oui' if avis.contrainte_mobilite else 'Non',
            avis.format_prefere, avis.observation,
            avis.created_at.strftime('%d/%m/%Y %H:%M') if avis.created_at else '',
        ])
        _style_data_row(ws, i, len(headers), alternate=(i % 2 == 0))

    _auto_width(ws)
    response = _excel_response(f"avis_{timezone.now().strftime('%Y%m%d')}.xlsx")
    wb.save(response)
    return response




# ── EXPORT 3 : Avis d'une session Excel ──────────────────────────────────────

@staff_member_required
def export_session_excel(request, pk):
    session = get_object_or_404(SessionPresentielle.objects.select_related('formation'), pk=pk)
    avis_list = Avis.objects.filter(session=session).order_by('nom')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avis"

    ws.merge_cells('A1:I1')
    ws['A1'] = f"Session : {session.formation.libelle}"
    ws['A1'].font = Font(bold=True, size=13, color='1F4E79')
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Date : {session.date.strftime('%d/%m/%Y')}  |  Lieu : {session.lieu}  |  Formateur : {session.formateur}"
    ws['A2'].font = Font(size=10, color='666666')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 28

    headers = [
        'Nom', 'Prénom', 'Structure', 'Fonction', 'Email', 'Téléphone',
        'Zone géographique', 'Format préféré', 'Contrainte mobilité',
        'Motivations', 'Observation', 'Date soumission',
    ]
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, 4, len(headers))

    for i, avis in enumerate(avis_list, start=5):
        ws.append([
            avis.nom, avis.prenom, avis.structure, avis.fonction,
            avis.email, avis.telephone, avis.zone_geographique,
            avis.format_prefere,
            'Oui' if avis.contrainte_mobilite else 'Non',
            avis.motivations, avis.observation,
            avis.created_at.strftime('%d/%m/%Y %H:%M') if avis.created_at else '',
        ])
        _style_data_row(ws, i, len(headers), alternate=(i % 2 == 0))

    _auto_width(ws)
    nom = session.formation.libelle[:30].replace(' ', '_')
    response = _excel_response(f"session_{nom}_{session.date.strftime('%Y%m%d')}.xlsx")
    wb.save(response)
    return response


# ── EXPORT 4 : Fiche session PDF ──────────────────────────────────────────────

@staff_member_required
def export_session_pdf(request, pk):
    session = get_object_or_404(SessionPresentielle.objects.select_related('formation'), pk=pk)
    avis_list = list(Avis.objects.filter(session=session).order_by('nom'))

    response = _pdf_response(
        f"session_{session.formation.libelle[:20]}_{session.date.strftime('%Y%m%d')}.pdf"
    )
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = _pdf_styles()
    elems = []

    elems.append(Paragraph("Bediakon Formation", styles['sous_titre']))
    elems.append(Paragraph(f"Fiche de session — {session.formation.libelle}", styles['titre']))
    elems.append(Spacer(1, 0.3*cm))

    info_data = [
        ['Date', session.date.strftime('%d/%m/%Y')],
        ['Lieu', session.lieu],
        ['Formateur', session.formateur],
        ['Capacité', f"{session.capacite} places"],
        ['Statut', session.get_statut_display()],
        ['Avis collectés', str(len(avis_list))],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (0, -1), PRIMARY),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, LIGHT_BLUE]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems.append(info_table)
    elems.append(Spacer(1, 0.6*cm))

    elems.append(Paragraph("Avis reçus", styles['section']))
    p_headers = ['#', 'Nom & Prénom', 'Structure', 'Email', 'Format préféré']
    p_data = [p_headers]
    for idx, avis in enumerate(avis_list, start=1):
        p_data.append([
            str(idx),
            f"{avis.nom} {avis.prenom}",
            avis.structure,
            avis.email,
            avis.format_prefere or '—',
        ])

    p_table = Table(p_data, colWidths=[0.8*cm, 4.5*cm, 4.5*cm, 5*cm, 1.5*cm])
    p_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elems.append(p_table)

    elems.append(Spacer(1, 1*cm))
    elems.append(Paragraph(
        f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}",
        styles['corps']
    ))

    doc.build(elems)
    response.write(buffer.getvalue())
    return response


# ── EXPORT 5 : Rapport analytique PDF ────────────────────────────────────────

@staff_member_required
def export_rapport_pdf(request):
    response = _pdf_response(f"rapport_analytique_{timezone.now().strftime('%Y%m%d')}.pdf")
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = _pdf_styles()
    elems = []

    elems.append(Paragraph("Bediakon Formation", styles['sous_titre']))
    elems.append(Paragraph("Rapport analytique des avis", styles['titre']))
    elems.append(Paragraph(
        f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}",
        styles['sous_titre']
    ))
    elems.append(Spacer(1, 0.4*cm))

    # Chiffres clés
    nb_formations = Formation.objects.filter(archive=False).count()
    nb_avis = Avis.objects.count()
    nb_sessions = SessionPresentielle.objects.count()

    elems.append(Paragraph("Chiffres clés", styles['section']))
    kpi_data = [
        ['Formations actives', 'Avis collectés', 'Sessions'],
        [str(nb_formations), str(nb_avis), str(nb_sessions)],
    ]
    kpi_table = Table(kpi_data, colWidths=[5.5*cm, 5.5*cm, 5*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, 1), 20),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elems.append(kpi_table)
    elems.append(Spacer(1, 0.6*cm))

    # Classement formations
    elems.append(Paragraph("Formations les plus demandées", styles['section']))
    formations = Formation.objects.filter(archive=False).annotate(
        nb=Count('sessions__avis')
    ).order_by('-nb')

    f_data = [['Formation', 'Avis collectés']]
    for f in formations:
        f_data.append([f.libelle, str(f.nb)])

    f_table = Table(f_data, colWidths=[13*cm, 3*cm])
    f_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems.append(f_table)
    elems.append(Spacer(1, 0.6*cm))

    # Répartition géographique
    elems.append(Paragraph("Répartition géographique", styles['section']))
    geo = (
        Avis.objects
        .exclude(zone_geographique='')
        .values('zone_geographique')
        .annotate(total=Count('id'))
        .order_by('-total')
    )
    geo_data = [['Zone géographique', 'Avis']]
    for g in geo:
        geo_data.append([g['zone_geographique'], str(g['total'])])
    if len(geo_data) == 1:
        geo_data.append(['Aucune donnée', '—'])

    geo_table = Table(geo_data, colWidths=[12*cm, 4*cm])
    geo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_BLUE]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems.append(geo_table)

    doc.build(elems)
    response.write(buffer.getvalue())
    return response
